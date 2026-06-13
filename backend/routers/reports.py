"""Reports: aggregated submission data for admin and per-student."""
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
from auth import get_current_user
from database import get_db

router = APIRouter()


def _fmt_time(secs):
    if secs is None:
        return ""
    return f"{secs // 60}m {secs % 60}s"


def _row(s: models.Submission) -> dict:
    return {
        "submission_id": s.id,
        "student_name": s.user.full_name or s.user.username if s.user else "—",
        "student_username": s.user.username if s.user else "—",
        "student_email": s.user.email if s.user else "—",
        "problem_title": s.problem.title if s.problem else "—",
        "mode": s.problem.mode.value if s.problem else "—",
        "status": s.status,
        "score": s.score,
        "time_taken": s.time_taken,
        "tab_switches": s.tab_switches,
        "test_cases_passed": s.test_cases_passed,
        "test_cases_total": s.test_cases_total,
        "submitted_at": s.submitted_at.isoformat() + ("" if s.submitted_at.tzinfo else "Z"),
    }


@router.get("")
def get_reports(
    mode: Optional[str] = None,      # practice | test
    student_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    query = db.query(models.Submission).join(models.Problem)

    # Students only see their own reports
    if current_user.role != "admin":
        query = query.filter(models.Submission.user_id == current_user.id)
    elif student_id:
        query = query.filter(models.Submission.user_id == student_id)

    if mode:
        query = query.filter(models.Problem.mode == mode)

    subs = query.order_by(models.Submission.submitted_at.desc()).all()
    return [_row(s) for s in subs]


@router.get("/export")
def export_reports(
    mode: Optional[str] = None,
    student_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Download submissions as a styled .xlsx (admin = all students, student = own)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    query = db.query(models.Submission).join(models.Problem)
    if current_user.role != "admin":
        query = query.filter(models.Submission.user_id == current_user.id)
    elif student_id:
        query = query.filter(models.Submission.user_id == student_id)
    if mode:
        query = query.filter(models.Problem.mode == mode)
    subs = query.order_by(models.Submission.submitted_at.desc()).all()

    is_admin = current_user.role == "admin"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submissions"

    headers = (["#", "Student", "Username", "Email"] if is_admin else ["#"]) + [
        "Problem", "Mode", "Status", "Score (%)", "Test Cases Passed",
        "Test Cases Total", "Time", "Tab Switches", "Submitted At",
    ]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="5C31FF")
    head_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")

    for i, s in enumerate(subs, 1):
        prob = s.problem.title if s.problem else "—"
        mode_v = s.problem.mode.value if s.problem else "—"
        submitted = s.submitted_at.strftime("%Y-%m-%d %H:%M") if s.submitted_at else ""
        base = [i]
        if is_admin:
            base += [
                (s.user.full_name or s.user.username) if s.user else "—",
                s.user.username if s.user else "—",
                s.user.email if s.user else "—",
            ]
        ws.append(base + [
            prob, mode_v, s.status, s.score,
            s.test_cases_passed, s.test_cases_total,
            _fmt_time(s.time_taken), s.tab_switches or 0, submitted,
        ])

    # widths
    widths = ([4, 22, 16, 26] if is_admin else [4]) + [26, 10, 16, 10, 16, 16, 10, 12, 18]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "codeforge_gradebook.xlsx" if is_admin else "codeforge_transcript.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{submission_id}")
def get_report_detail(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    s = (
        db.query(models.Submission)
        .filter(models.Submission.id == submission_id)
        .first()
    )
    if not s:
        from fastapi import HTTPException
        raise HTTPException(404, "Not found")
    if current_user.role != "admin" and s.user_id != current_user.id:
        from fastapi import HTTPException
        raise HTTPException(403, "Forbidden")

    results = [
        {
            "test_case_id": r.test_case_id,
            "status": r.status,
            "actual_output": r.actual_output,
            "execution_time": r.execution_time,
        }
        for r in s.results
    ]

    return {
        **_row(s),
        "code": s.code,
        "feedback": s.feedback,
        "results": results,
    }
